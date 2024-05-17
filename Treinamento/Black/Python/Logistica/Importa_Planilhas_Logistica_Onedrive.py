from bs4 import BeautifulSoup
import os
import pandas as pd
import pyodbc
import sqlalchemy
from openpyxl import load_workbook

from datetime import timedelta, date


# Definir conexao 0 = PowerBiV2 / 1 = Datawarehouse
opcao = 0

configuracoes = pd.read_csv("C:/Python/database.cfg", header = 0, delimiter=";")

server = configuracoes['IP'].values[opcao]
database = configuracoes['Banco'].values[opcao]
username = configuracoes['Usuario'].values[opcao]
password = configuracoes['Senha'].values[opcao]


sql_string = 'mssql+pyodbc://' + username + ':' + password + '@' + server + '/' + database + '?driver=ODBC+Driver+13+for+SQL+Server'

cnxn = pyodbc.connect(
    'DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
cursor = cnxn.cursor()

# Conexao com o servidor SQL
engine = sqlalchemy.create_engine(sql_string)
engine.connect()

# Data de hoje
# hoje = date.today().strftime('%d/%m/%Y')
hoje = date.today()
hojef = hoje.strftime('%d/%m/%Y')


# Busca os dados da planilha para pesquisar nos sites
desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'OneDrive - Mais Proxima\Importacao_Dados')
arquivo = '\LogES\coletas2020.xlsx'
desktop = desktop_path + arquivo


# IMPORTAÇAO PLANILHA DE COLETAS

try:
    dados = pd.read_excel(desktop, sheet_name='Coletas')
    qtd_reg = str(len(dados))
    del dados['OBS']
    dados = dados.astype(str)

    dados.to_sql(
        name='Tabela_Onedrive_Coletas_Temp',  # database table name
        con=engine,
        if_exists='replace',
        index=False,
    )

    # Grava na planilha que o registro foi importado para o SQL
    book = load_workbook(desktop)
    ws = book.worksheets[0]
    lin = 2
    col = 12
    texto = 'Importado para SQL em ' + str(hojef)

    for items in ws:
        ws.cell(row=lin, column=col).value = texto

        if lin <= len(dados):
            lin = lin + 1
    book.save(desktop)
    book.close()
    print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)

except:
    print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')


# IMPORTAÇAO PLANILHA DE COTACAO
arquivo = '\LogES\cotações2020.xlsx'
desktop = desktop_path + arquivo

try:
    dados = pd.read_excel(desktop, sheet_name='Cotacao')
    qtd_reg = str(len(dados))
    del dados['OBS']
    dados = dados.astype(str)

    dados.to_sql(
        name='Tabela_Onedrive_Cotacao_Temp',  # database table name
        con=engine,
        if_exists='replace',
        index=False,
    )

    # Grava na planilha que o registro foi importado para o SQL
    book = load_workbook(desktop)
    ws = book.worksheets[0]
    lin = 2
    col = 9
    texto = 'Importado para SQL em ' + str(hojef)

    for items in ws:
        ws.cell(row=lin, column=col).value = texto

        if lin <= len(dados):
            lin = lin + 1
    book.save(desktop)
    book.close()
    print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)


except:
    print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')




# IMPORTAÇAO PLANILHA NFS E ESTOQUE DIA
arquivo = '\LogES\estoque_nfs_expedidas_dia.xlsx'
desktop = desktop_path + arquivo

try:
    dados = pd.read_excel(desktop, sheet_name='Expedicao')
    qtd_reg = str(len(dados))
    del dados['OBS']
    dados.fillna(0)
    dados = dados.astype(str)

    dados.to_sql(
        name='Tabela_Onedrive_Expedicoes_Dia_Temp',  # database table name
        con=engine,
        if_exists='replace',
        index=False,
    )

    # Grava na planilha que o registro foi importado para o SQL
    book = load_workbook(desktop)
    ws = book.worksheets[0]
    lin = 2
    col = 10
    texto = 'Importado para SQL em ' + str(hojef)

    for items in ws:
        ws.cell(row=lin, column=col).value = texto

        if lin <= len(dados):
            lin = lin + 1
    book.save(desktop)
    book.close()
    print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)


except:
    print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')




# IMPORTAÇAO PLANILHA EXPEDICOES
arquivo = '\LogES\expedições_2020.xlsx'
desktop = desktop_path + arquivo

try:
    dados = pd.read_excel(desktop, sheet_name='Base_expedicoes')
    qtd_reg = str(len(dados))
    del dados['OBS']

    dados = dados.astype(str)

    dados.to_sql(
        name='Tabela_Onedrive_Nfs_Expedidas_Temp',  # database table name
        con=engine,
        if_exists='replace',
        index=False,
    )

    # Grava na planilha que o registro foi importado para o SQL
    book = load_workbook(desktop)
    ws = book.worksheets[0]
    lin = 2
    col = 8
    texto = 'Importado para SQL em ' + str(hojef)

    for items in ws:
        ws.cell(row=lin, column=col).value = texto

        if lin <= len(dados):
            lin = lin + 1
    book.save(desktop)
    book.close()
    print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)


except:
    print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')





# IMPORTAÇAO PLANILHA NFS PENDENTES DE EXPEDICAO
arquivo = '\LogES\pendentes_expedição.xlsx'
desktop = desktop_path + arquivo

try:
    dados = pd.read_excel(desktop, sheet_name='NF_Pendentes')
    qtd_reg = str(len(dados))
    del dados['OBS']
    dados = dados.astype(str)

    dados.to_sql(
        name='Tabela_Onedrive_NF_Pendentes_Temp',  # database table name
        con=engine,
        if_exists='replace',
        index=False,
    )

    # Grava na planilha que o registro foi importado para o SQL
    book = load_workbook(desktop)
    ws = book.worksheets[0]
    lin = 2
    col = 3
    texto = 'Importado para SQL em ' + str(hojef)

    for items in ws:
        ws.cell(row=lin, column=col).value = texto

        if lin <= len(dados):
            lin = lin + 1
    book.save(desktop)
    book.close()
    print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)


except:
    print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')
