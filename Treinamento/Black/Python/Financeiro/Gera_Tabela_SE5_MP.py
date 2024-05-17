# coding=utf8

import os
import pandas as pd
from openpyxl import load_workbook
import time
import pyodbc
import sqlalchemy
from datetime import timedelta, date

def diff_days(date1, date2):
    d1 = datetime.strptime(date1, "%d-%m-%Y")
    d2 = datetime.strptime(date2, "%d-%m-%Y")
    return abs((date2 - date1).days)


# Definir conexao 0 = PowerBiV2 / 1 = Datawarehouse
opcao = 1

configuracoes = pd.read_csv("C:/Python/database.cfg", header = 0, delimiter=";")

server = configuracoes['IP'].values[opcao]
database = configuracoes['Banco'].values[opcao]
username = configuracoes['Usuario'].values[opcao]
password = configuracoes['Senha'].values[opcao]


sql_string = 'mssql+pyodbc://' + username + ':' + password + '@' + server + '/' + database + '?driver=ODBC+Driver+13+for+SQL+Server'

cnxn = pyodbc.connect(
    'DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
cursor = cnxn.cursor()

# Conexao com o servidor SQL
engine = sqlalchemy.create_engine(sql_string)
engine.connect()


# Data de hoje
hoje = date.today().strftime('%d/%m/%Y')

# Busca os dados da planilha para pesquisar nos sites
# desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
desktop_path = 'P:\Financeiro'
arquivo = '\Base.xlsx'
desktop = desktop_path + arquivo
dados = pd.read_excel(desktop, sheet_name='Base')
dados_protheus = pd.read_sql_table('v_Mapa_FC_Real', engine)
dados_protheus.sort_values(by='E5_DATA', ascending=True)

#Limpando Planilha
book = load_workbook(desktop)
ws = book.worksheets[0]

cont = 1
qtd_reg = len(dados)
range_plan = 'A2:DD'+str(qtd_reg)

print('Limpando planilha excel Base.xlsx')

for row in ws[range_plan]:
  for cell in row:
    cell.value = None

book.save(desktop)
book.close()

linha = 2

# Gravando dados na planílha
book = load_workbook(desktop)
ws = book.worksheets[0]

print("Gravando dados na planilha")
for index, row in dados_protheus.iterrows():
    ws.cell(row=linha, column=1).value = str(dados_protheus['E5_FILIAL'][index])
    ws.cell(row=linha, column=2).value = str(dados_protheus['E5_DATA'][index])
    ws.cell(row=linha, column=3).value = str(dados_protheus['E5_TIPO'][index])
    ws.cell(row=linha, column=4).value = str(dados_protheus['E5_MOEDA'][index])
    ws.cell(row=linha, column=5).value = str(dados_protheus['E5_VALOR'][index])
    ws.cell(row=linha, column=6).value = str(dados_protheus['E5_NATUREZ'][index])
    ws.cell(row=linha, column=7).value = str(dados_protheus['E5_BANCO'][index])
    ws.cell(row=linha, column=8).value = str(dados_protheus['E5_AGENCIA'][index])
    ws.cell(row=linha, column=9).value = str(dados_protheus['E5_CONTA'][index])
    ws.cell(row=linha, column=10).value = str(dados_protheus['E5_NUMCHEQ'][index])
    ws.cell(row=linha, column=11).value = str(dados_protheus['E5_DOCUMEN'][index])
    ws.cell(row=linha, column=12).value = str(dados_protheus['E5_VENCTO'][index])
    ws.cell(row=linha, column=13).value = str(dados_protheus['E5_RECPAG'][index])
    ws.cell(row=linha, column=14).value = str(dados_protheus['E5_BENEF'][index])
    ws.cell(row=linha, column=15).value = str(dados_protheus['E5_HISTOR'][index])
    ws.cell(row=linha, column=16).value = str(dados_protheus['E5_TIPODOC'][index])
    ws.cell(row=linha, column=17).value = str(dados_protheus['E5_VLMOED2'][index])
    ws.cell(row=linha, column=18).value = str(dados_protheus['E5_LA'][index])
    ws.cell(row=linha, column=19).value = str(dados_protheus['E5_SITUACA'][index])
    ws.cell(row=linha, column=20).value = str(dados_protheus['E5_LOTE'][index])
    ws.cell(row=linha, column=21).value = str(dados_protheus['E5_PREFIXO'][index])
    ws.cell(row=linha, column=22).value = str(dados_protheus['E5_NUMERO'][index])
    ws.cell(row=linha, column=23).value = str(dados_protheus['E5_PARCELA'][index])
    ws.cell(row=linha, column=24).value = str(dados_protheus['E5_CLIFOR'][index])
    ws.cell(row=linha, column=25).value = str(dados_protheus['E5_LOJA'][index])
    ws.cell(row=linha, column=26).value = str(dados_protheus['E5_DTDIGIT'][index])
    ws.cell(row=linha, column=27).value = str(dados_protheus['E5_TIPOLAN'][index])
    ws.cell(row=linha, column=28).value = str(dados_protheus['E5_DEBITO'][index])
    ws.cell(row=linha, column=29).value = str(dados_protheus['E5_CREDITO'][index])
    ws.cell(row=linha, column=30).value = str(dados_protheus['E5_MOTBX'][index])
    ws.cell(row=linha, column=31).value = str(dados_protheus['E5_RATEIO'][index])
    ws.cell(row=linha, column=32).value = str(dados_protheus['E5_RECONC'][index])
    ws.cell(row=linha, column=33).value = str(dados_protheus['E5_SEQ'][index])
    ws.cell(row=linha, column=34).value = str(dados_protheus['E5_DTDISPO'][index])
    ws.cell(row=linha, column=35).value = str(dados_protheus['E5_CCD'][index])
    ws.cell(row=linha, column=36).value = str(dados_protheus['E5_CCC'][index])
    ws.cell(row=linha, column=37).value = str(dados_protheus['E5_OK'][index])
    ws.cell(row=linha, column=38).value = str(dados_protheus['E5_ARQRAT'][index])
    ws.cell(row=linha, column=39).value = str(dados_protheus['E5_IDENTEE'][index])
    ws.cell(row=linha, column=40).value = str(dados_protheus['E5_ORDREC'][index])
    ws.cell(row=linha, column=41).value = str(dados_protheus['E5_FILORIG'][index])
    ws.cell(row=linha, column=42).value = str(dados_protheus['E5_ARQCNAB'][index])
    ws.cell(row=linha, column=43).value = str(dados_protheus['E5_VLJUROS'][index])
    ws.cell(row=linha, column=44).value = str(dados_protheus['E5_VLMULTA'][index])
    ws.cell(row=linha, column=45).value = str(dados_protheus['E5_VLCORRE'][index])
    ws.cell(row=linha, column=46).value = str(dados_protheus['E5_VLDESCO'][index])
    ws.cell(row=linha, column=47).value = str(dados_protheus['E5_CNABOC'][index])
    ws.cell(row=linha, column=48).value = str(dados_protheus['E5_SITUA'][index])
    ws.cell(row=linha, column=49).value = str(dados_protheus['E5_ITEMD'][index])
    ws.cell(row=linha, column=50).value = str(dados_protheus['E5_ITEMC'][index])
    ws.cell(row=linha, column=51).value = str(dados_protheus['E5_CLVLDB'][index])
    ws.cell(row=linha, column=52).value = str(dados_protheus['E5_CLVLCR'][index])
    ws.cell(row=linha, column=53).value = str(dados_protheus['E5_PROJPMS'][index])
    ws.cell(row=linha, column=54).value = str(dados_protheus['E5_EDTPMS'][index])
    ws.cell(row=linha, column=55).value = str(dados_protheus['E5_TASKPMS'][index])
    ws.cell(row=linha, column=56).value = str(dados_protheus['E5_MODSPB'][index])
    ws.cell(row=linha, column=57).value = str(dados_protheus['E5_FATURA'][index])
    ws.cell(row=linha, column=58).value = str(dados_protheus['E5_TXMOEDA'][index])
    ws.cell(row=linha, column=59).value = str(dados_protheus['E5_CODORCA'][index])
    ws.cell(row=linha, column=60).value = str(dados_protheus['E5_FATPREF'][index])
    ws.cell(row=linha, column=61).value = str(dados_protheus['E5_SITCOB'][index])
    ws.cell(row=linha, column=62).value = str(dados_protheus['E5_FORNADT'][index])
    ws.cell(row=linha, column=63).value = str(dados_protheus['E5_LOJAADT'][index])
    ws.cell(row=linha, column=64).value = str(dados_protheus['E5_CLIENTE'][index])
    ws.cell(row=linha, column=65).value = str(dados_protheus['E5_FORNECE'][index])
    ws.cell(row=linha, column=66).value = str(dados_protheus['E5_SERREC'][index])
    ws.cell(row=linha, column=67).value = str(dados_protheus['E5_OPERAD'][index])
    ws.cell(row=linha, column=68).value = str(dados_protheus['E5_MOVCX'][index])
    ws.cell(row=linha, column=69).value = str(dados_protheus['E5_KEY'][index])
    ws.cell(row=linha, column=70).value = str(dados_protheus['E5_MULTNAT'][index])
    ws.cell(row=linha, column=71).value = str(dados_protheus['E5_AGLIMP'][index])
    ws.cell(row=linha, column=72).value = str(dados_protheus['E5_VLACRES'][index])
    ws.cell(row=linha, column=73).value = str(dados_protheus['E5_VLDECRE'][index])
    ws.cell(row=linha, column=74).value = str(dados_protheus['E5_VRETPIS'][index])
    ws.cell(row=linha, column=75).value = str(dados_protheus['E5_VRETCOF'][index])
    ws.cell(row=linha, column=76).value = str(dados_protheus['E5_VRETCSL'][index])
    ws.cell(row=linha, column=77).value = str(dados_protheus['E5_PRETPIS'][index])
    ws.cell(row=linha, column=78).value = str(dados_protheus['E5_PRETCOF'][index])
    ws.cell(row=linha, column=79).value = str(dados_protheus['E5_PRETCSL'][index])
    ws.cell(row=linha, column=80).value = str(dados_protheus['E5_AUTBCO'][index])
    ws.cell(row=linha, column=81).value = str(dados_protheus['E5_PRETIRF'][index])
    ws.cell(row=linha, column=82).value = str(dados_protheus['E5_VRETIRF'][index])
    ws.cell(row=linha, column=83).value = str(dados_protheus['E5_VRETISS'][index])
    ws.cell(row=linha, column=84).value = str(dados_protheus['E5_NUMMOV'][index])
    ws.cell(row=linha, column=85).value = str(dados_protheus['E5_DIACTB'][index])
    ws.cell(row=linha, column=86).value = str(dados_protheus['E5_NODIA'][index])
    ws.cell(row=linha, column=87).value = str(dados_protheus['E5_BASEIRF'][index])
    ws.cell(row=linha, column=88).value = str(dados_protheus['E5_PROCTRA'][index])
    ws.cell(row=linha, column=89).value = str(dados_protheus['E5_USERLGI'][index])
    ws.cell(row=linha, column=90).value = str(dados_protheus['E5_USERLGA'][index])
    ws.cell(row=linha, column=91).value = str(dados_protheus['E5_ORIGEM'][index])
    ws.cell(row=linha, column=92).value = str(dados_protheus['E5_FORMAPG'][index])
    ws.cell(row=linha, column=93).value = str(dados_protheus['E5_TPDESC'][index])
    ws.cell(row=linha, column=94).value = str(dados_protheus['E5_PRISS'][index])
    ws.cell(row=linha, column=95).value = str(dados_protheus['E5_PRINSS'][index])
    ws.cell(row=linha, column=96).value = str(dados_protheus['E5_FLDMED'][index])
    ws.cell(row=linha, column=97).value = str(dados_protheus['E5_CGC'][index])
    ws.cell(row=linha, column=98).value = str(dados_protheus['E5_IDMOVI'][index])
    ws.cell(row=linha, column=99).value = str(dados_protheus['E5_XOUTCON'][index])
    ws.cell(row=linha, column=100).value = str(dados_protheus['E5_DTCANBX'][index])
    ws.cell(row=linha, column=101).value = str(dados_protheus['E5_PRETINS'][index])
    ws.cell(row=linha, column=102).value = str(dados_protheus['E5_VRETINS'][index])
    ws.cell(row=linha, column=103).value = str(dados_protheus['E5_CCUSTO'][index])
    ws.cell(row=linha, column=104).value = str(dados_protheus['E5_SEQCON'][index])
    ws.cell(row=linha, column=105).value = str(dados_protheus['E5_SDOCREC'][index])
    ws.cell(row=linha, column=106).value = str(dados_protheus['E5_MOVFKS'][index])
    ws.cell(row=linha, column=107).value = str(dados_protheus['E5_IDORIG'][index])
    ws.cell(row=linha, column=108).value = str(dados_protheus['E5_TABORI'][index])

    linha = linha + 1

book.save(desktop)
book.close()