from bs4 import BeautifulSoup
import os
import pandas as pd
import pyodbc
import sqlalchemy
from openpyxl import load_workbook
import re
import time
from datetime import timedelta, date


# Definir conexao 0 = PowerBiV2 / 1 = Datawarehouse
opcao = 0

configuracoes = pd.read_csv("C:/Python/database.cfg", header = 0, delimiter=";")

server = configuracoes['IP'].values[opcao]
database = configuracoes['Banco'].values[opcao]
username = configuracoes['Usuario'].values[opcao]
password = configuracoes['Senha'].values[opcao]


sql_string = 'mssql+pyodbc://' + username + ':' + password + '@' + server + '/' + database + '?driver=ODBC+Driver+13+for+SQL+Server'

cnxn = pyodbc.connect(
    'DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
cursor = cnxn.cursor()

# Conexao com o servidor SQL
engine = sqlalchemy.create_engine(sql_string)
engine.connect()



# Data de hoje
hoje = date.today().strftime('%d/%m/%Y')


# Busca os dados da planilha para pesquisar nos sites
desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'OneDrive - Mais Proxima\Importacao_Dados')
arquivo = '\Logistica_MP.xlsx'
desktop = desktop_path + arquivo


# IMPORTANDO COLETA DE FRETES
try:
    dados = pd.read_excel(desktop, sheet_name='Coletas')
    del dados['OBS']
    dados = dados.astype(str)


    dados.to_sql(
        name='Tabela_Frete_Coleta_Temp', # database table name
        con=engine,
        if_exists='replace',
        index=False,
    )

    # Grava na planilha que o registro foi importado para o SQL
    book = load_workbook(desktop)
    ws = book.worksheets[1]
    ws0 = ws

    hoje = date.today()
    hojef = hoje.strftime('%d/%m/%Y')
    texto = 'Importado para SQL em ' + str(hojef)
    lin = 2
    cell = 12

    for items in ws:
        ws.cell(row=lin, column=cell).value = texto

        if lin <= len(dados):
            lin = lin + 1
    book.save(desktop)
    book.close()

    # Executando PROCEDURE no banco SQL
    print('Executando PROCEDURE FreteColeta no banco SQL')
    cursor.execute("EXEC FreteColeta")
    cursor.commit()
    totalregistros = str(len(dados)) + ' registros importados'
    print('Importacao Coletas realizada com sucesso...', totalregistros)

except:
    print('A planilha '+ arquivo +' esta aberta, importacao nao realizada...')



# IMPORTANDO COTAÇOES DE FRETE
try:
    dados = pd.read_excel(desktop, sheet_name='Cotacao')
    del dados['OBS']
    dados = dados.astype(str)

    dados.to_sql(
        name='Tabela_Frete_Cotacao_Temp', # database table name
        con=engine,
        if_exists='replace',
        index=False,
    )

    # Grava na planilha que o registro foi importado para o SQL
    book = load_workbook(desktop)
    ws = book.worksheets[0]
    ws1 = ws

    hoje = date.today()
    hojef = hoje.strftime('%d/%m/%Y')
    texto = 'Importado para SQL em ' + str(hojef)
    lin = 2
    cell = 9

    for items in ws:
        ws.cell(row=lin, column=cell).value = texto

        if lin <= len(dados):
            lin = lin + 1
    book.save(desktop)
    book.close()

    # Executando PROCEDURE no banco SQL
    print('Executando PROCEDURE FreteCotado no banco SQL')
    cursor.execute("EXEC FreteCotado")
    cursor.commit()

    totalregistros = str(len(dados)) + ' registros importados'
    print('Importacao Cotacoes realizada com sucesso...', totalregistros)

except:
    print('A planilha '+ arquivo +' esta aberta, importacao nao realizada...')




# IMPORTANDO EXPEDIÇOES
try:
    dados = pd.read_excel(desktop, sheet_name='Expedicao')
    del dados['OBS']
    dados = dados.astype(str)

    dados.to_sql(
        name='Tabela_Operacao_Logistica_Temp', # database table name
        con=engine,
        if_exists='replace',
        index=False,
    )

    # Grava na planilha que o registro foi importado para o SQL
    book = load_workbook(desktop)
    ws = book.worksheets[2]
    ws2 = ws

    hoje = date.today()
    hojef = hoje.strftime('%d/%m/%Y')
    texto = 'Importado para SQL em ' + str(hojef)
    lin = 2
    cell = 11

    for items in ws:
        ws.cell(row=lin, column=cell).value = texto

        if lin <= len(dados):
            lin = lin + 1
    book.save(desktop)
    book.close()

    totalregistros = str(len(dados)) + ' registros importados'
    print('Importacao Expedicoes realizada com sucesso...', totalregistros)

except:
    print('A planilha '+ arquivo +' esta aberta, importacao nao realizada...')



# IMPORTANDO RECEBIMENTO
try:
    dados = pd.read_excel(desktop, sheet_name='Recebimentos')
    del dados['OBS']
    dados = dados.astype(str)

    dados.to_sql(
        name='Tabela_Operacao_Logistica3_Temp', # database table name
        con=engine,
        if_exists='replace',
        index=False,
    )

    # Grava na planilha que o registro foi importado para o SQL
    book = load_workbook(desktop)
    ws = book.worksheets[3]
    ws3 = ws

    hoje = date.today()
    hojef = hoje.strftime('%d/%m/%Y')
    texto = 'Importado para SQL em ' + str(hojef)
    lin = 2
    cell = 13

    for items in ws:
        ws.cell(row=lin, column=cell).value = texto

        if lin <= len(dados):
            lin = lin + 1
    book.save(desktop)
    book.close()

    totalregistros = str(len(dados)) + ' registros importados'
    print('Importacao Recebimento realizada com sucesso...', totalregistros)

except:
    print('A planilha '+ arquivo +' esta aberta, importacao nao realizada...')



# IMPORTANDO NFS PENDENTES DE EXPEDIÇAO
try:
    dados = pd.read_excel(desktop, sheet_name='NFS PENDENTES EXPEDIÇÃO POR DIA')
    # del dados['OBS']
    dados = dados.astype(str)

    dados.to_sql(
        name='Tabela_Operacao_Logistica2_Temp', # database table name
        con=engine,
        if_exists='replace',
        index=False,
    )

    # Grava na planilha que o registro foi importado para o SQL
    book = load_workbook(desktop)
    ws = book.worksheets[4]
    ws4 = ws

    hoje = date.today()
    hojef = hoje.strftime('%d/%m/%Y')
    texto = 'Importado para SQL em ' + str(hojef)
    lin = 2
    cell = 3

    for items in ws:
        ws.cell(row=lin, column=cell).value = texto

        if lin <= len(dados):
            lin = lin + 1
    book.save(desktop)
    book.close()

    totalregistros = str(len(dados)) + ' registros importados'
    print('Importacao NF´s Pendentes realizada com sucesso...', totalregistros)

except:
    print('A planilha '+ arquivo +' esta aberta, importacao nao realizada...')



# Executando PROCEDURE OperacaoLog no banco SQL
print('Executando PROCEDURE OperacaoLog no banco SQL')
cursor.execute("EXEC Datawarehouse..OperacaoLog")
cursor.commit()




