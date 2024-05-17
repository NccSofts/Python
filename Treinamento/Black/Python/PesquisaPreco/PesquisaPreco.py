from bs4 import BeautifulSoup
import os
import pandas as pd
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import  Keys
import pyautogui as pyg
import re
import time
from datetime import timedelta, date

def check_exists_by_xpath(xpath):
    try:
        browser.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True


browser = webdriver.Chrome(executable_path=r'C:/Python/Selenium/chromedriver.exe')  # Optional argument, if not specified will search path
browser.maximize_window()
action = ActionChains(browser)

# Data de hoje
hoje = date.today().strftime('%d/%m/%Y')


# Busca os dados da planilha para pesquisar nos sites
desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
arquivo = '\Pesquisa_Preco_MP.xlsx'
desktop = desktop_path + arquivo
dados = pd.read_excel(desktop, sheet_name='ProdutosPesquisa')



cont = 0
linha = 2

for index, row in dados.iterrows():
    browser.get('https://www.buscape.com.br/')
    bar_procura = browser.find_element_by_id("main-search")
    bar_procura.send_keys(dados['CodigoFabricante'][index])
    bar_procura.send_keys(Keys.ENTER)

    #Checando botoes de produtos após procura
    chk_btn_compare = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/a[2]/div/button")
    if chk_btn_compare == True:
        btn_prod = browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/a[2]/div/button").text
        btn_compare = browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/a[2]/div/button")

    chk_btn_irloja = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div/div/a[2]/div/button")
    if chk_btn_irloja == True:
        btn_prod = browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div/div/a[2]/div/button").text

    if chk_btn_compare == False and chk_btn_irloja == False:
        browser.get('https://www.buscape.com.br/')
        bar_procura = browser.find_element_by_id("main-search")
        bar_procura.send_keys(dados['Produto'][index])
        bar_procura.send_keys(Keys.ENTER)

        chk_btn_compare = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/a[2]/div/button")
        if chk_btn_compare == True:
            btn_prod = browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/a[2]/div/button").text
            btn_compare = browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/a[2]/div/button")

        chk_btn_irloja = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div/div/a[2]/div/button")
        if chk_btn_irloja == True:
            btn_prod = browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div/div/a[2]/div/button").text

        if chk_btn_compare == False and chk_btn_irloja == False:
            btn_prod = 'PRODUTO NÃO LOCALIZADO'
            tipo_pesquisa = 'Pela Descrição do Produto'



    if btn_prod == 'COMPARE MAIS PREÇOS':
        browser.execute_script("arguments[0].click();", btn_compare)
        ActionChains(browser).send_keys(Keys.PAGE_DOWN).perform()
        time.sleep(0.5)
        ActionChains(browser).send_keys(Keys.PAGE_DOWN).perform()
        time.sleep(0.5)
        ActionChains(browser).send_keys(Keys.PAGE_UP).perform()
        ActionChains(browser).send_keys(Keys.PAGE_UP).perform()
        tipo_pesquisa = 'Pelo Código Fabricante'


        chk_menor_preco = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/div/div/ul/li[2]")
        if chk_menor_preco == True:
            btn_menor_preco = browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/div/div/ul/li[2]")
            browser.execute_script("arguments[0].click();", btn_menor_preco)

        chk_desc_prod = check_exists_by_xpath("//*[@id='app']/div/div[1]/div/div/div/div/section/div[2]/div[1]/h1")
        if chk_desc_prod == True:
            descr_prod = str(browser.find_element_by_xpath("//*[@id='app']/div/div[1]/div/div/div/div/section/div[2]/div[1]/h1").text)

        else:
            descr_prod = ''

        chk_loja1 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[1]/a/div[2]/img")
        if chk_loja1 == True:
            loja1 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[1]/a/div[2]/img").get_attribute('alt'))
        else:
            loja1 = ''

        chk_loja2 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[2]/a/div[2]/img")
        if chk_loja2 == True:
            loja2 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[2]/a/div[2]/img").get_attribute('alt'))
        else:
            loja2 = ''

        chk_loja3 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[3]/a/div[2]/img")
        if chk_loja3 == True:
            loja3 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[3]/a/div[2]/img").get_attribute('alt'))
        else:
            loja3 = ''

        chk_preco1 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[1]/a/div[4]/div[2]/span[2]/div/span")
        if chk_preco1 == True:
            preco1 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[1]/a/div[4]/div[2]/span[2]/div/span").text)
        else:
            preco1 = ''

        chk_preco2 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[2]/a/div[4]/div/span[2]/div/span")
        if chk_preco2 == True:
            preco2 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[2]/a/div[4]/div/span[2]/div/span").text)
        else:
            preco2 = ''

        chk_preco3 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[3]/a/div[4]/div/span[2]/div/span")
        if chk_preco3 == True:
            preco3 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[3]/a/div[4]/div/span[2]/div/span").text)
        else:
            preco3 = ''


        chk_parcela1 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[1]/a/div[5]/div/span[1]")
        if chk_parcela1 == True:
            parcelamento1 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[1]/a/div[5]/div/span[1]").text)
        else:
            parcelamento1 = ''

        chk_parcela2 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[2]/a/div[5]/div/span[1]")
        if chk_parcela2 == True:
            parcelamento2 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[2]/a/div[5]/div/span[1]").text)
        else:
            parcelamento2 = ''

        chk_parcela3 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[3]/a/div[5]/div/span[1]")
        if chk_parcela3 == True:
            parcelamento3 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[3]/a/div[5]/div/span[1]").text)
        else:
            parcelamento3 = ''


        chk_valor_parc1 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[1]/a/div[5]/div/span[2]/div/span")
        if chk_valor_parc1 == True:
            valor_parc1 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[1]/a/div[5]/div/span[2]/div/span").text)
        else:
            valor_parc1 = ''

        chk_valor_parc2 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[2]/a/div[5]/div/span[2]/div/span")
        if chk_valor_parc2 == True:
            valor_parc2 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[2]/a/div[5]/div/span[2]/div/span").text)
        else:
            valor_parc2 = ''

        chk_valor_parc3 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[3]/a/div[5]/div/span[2]/div/span")
        if chk_valor_parc3 == True:
            valor_parc3 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[3]/a/div[5]/div/span[2]/div/span").text)
        else:
            valor_parc3 = ''


        chk_total_prazo1 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[1]/a/div[5]/div/span[3]")
        if chk_total_prazo1 == True:
            total_parc1 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[1]/a/div[5]/div/span[3]").text)
        else:
            total_parc1 = ''

        chk_total_prazo2 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[2]/a/div[5]/div/span[3]")
        if chk_total_prazo2 == True:
            total_parc2 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[2]/a/div[5]/div/span[3]").text)
        else:
            total_parc2 = ''

        chk_total_prazo3 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[3]/a/div[5]/div/span[3]")
        if chk_total_prazo3 == True:
            total_parc3 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[1]/div/div/div/div[2]/li[3]/a/div[5]/div/span[3]").text)
        else:
            total_parc3 = ''

        chk_preco = check_exists_by_xpath("//*[@id='app']/div/div[1]/div/div/div/div/section/div[2]/div[3]/a/div/span[1]/span")
        if chk_preco == True:
            preco1 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[1]/div/div/div/div/section/div[2]/div[3]/a/div/span[1]/span").text)
            loja1 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[1]/div/div/div/div/section/div[2]/div[3]/a/span").text)
            parcelamento1 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[1]/div/div/div/div/section/div[2]/div[3]/a/div/span[2]/span").text)


        # Gravando dados na planílha

        book = load_workbook(desktop)
        ws = book.worksheets[0]
        ws.cell(row=linha, column=7).value = hoje
        ws.cell(row=linha, column=8).value = tipo_pesquisa
        ws.cell(row=linha, column=9).value = descr_prod

        #Loja 1
        ws.cell(row=linha, column=10).value = loja1
        ws.cell(row=linha, column=11).value = preco1
        ws.cell(row=linha, column=12).value = parcelamento1
        ws.cell(row=linha, column=13).value = valor_parc1
        ws.cell(row=linha, column=14).value = total_parc1

        #Loja 2
        ws.cell(row=linha, column=15).value = loja2
        ws.cell(row=linha, column=16).value = preco2
        ws.cell(row=linha, column=17).value = parcelamento2
        ws.cell(row=linha, column=18).value = valor_parc2
        ws.cell(row=linha, column=19).value = total_parc2

        #Loja 3
        ws.cell(row=linha, column=20).value = loja3
        ws.cell(row=linha, column=21).value = preco3
        ws.cell(row=linha, column=22).value = parcelamento3
        ws.cell(row=linha, column=23).value = valor_parc3
        ws.cell(row=linha, column=24).value = total_parc3

        book.save(desktop)
        book.close()

        cont = cont + 1
        linha = linha + 1


    if btn_prod == 'IR À LOJA':

        chk_menor_preco = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/section/div[3]/div/div/div/ul/li[2]/a")
        if chk_menor_preco == True:
            btn_menor_preco = browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/section/div[3]/div/div/div/ul/li[2]/a")
            browser.execute_script("arguments[0].click();", btn_menor_preco)

        chk_preco1 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div/div/a[1]/div[4]/div/div/span[1]/span/span[2]")
        if chk_preco1 == True:
            descr_prod = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div/div/a[1]/div[3]").text)
            preco1 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div/div/a[1]/div[4]/div/div/span[1]/span/span[2]").text)
            loja1 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div/div/a[1]/div[4]/div/span/span").text)
            parcelamento1 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[1]/div/div/div/a[1]/div[4]/div/div/span[2]/span").text)
        else:
            descr_prod = ''
            preco1 = ''
            loja1 = ''
            parcelamento1 = ''


        chk_preco2 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[2]/div/div/div/a[1]/div[4]/div/div/span[1]/span/span[2]")
        if chk_preco2 == True:
            preco2 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[2]/div/div/div/a[1]/div[4]/div/div/span[1]/span/span[2]").text)
            loja2 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[2]/div/div/div/a[1]/div[4]/div/span/span").text)
            parcelamento2 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[2]/div/div/div/a[1]/div[4]/div/div/span[2]/span").text)
        else:
            descr_prod = ''
            preco1 = ''
            loja1 = ''
            parcelamento2 = ''


        chk_preco3 = check_exists_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div/a[1]/div[4]/div/div/span[1]/span/span[2]")
        if chk_preco3 == True:
            preco3 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div/a[1]/div[4]/div/div/span[1]/span/span[2]").text)
            loja3 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div/a[1]/div[4]/div/span/span").text)
            parcelamento3 = str(browser.find_element_by_xpath("//*[@id='app']/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div/a[1]/div[4]/div/div/span[2]/span").text)
        else:
            descr_prod = ''
            preco3 = ''
            loja3 = ''
            parcelamento3 = ''

        # Gravando dados na planílha

        book = load_workbook(desktop)
        ws = book.worksheets[0]
        ws.cell(row=linha, column=7).value = hoje
        ws.cell(row=linha, column=8).value = tipo_pesquisa
        ws.cell(row=linha, column=9).value = descr_prod

        # Preço Lojas
        ws.cell(row=linha, column=10).value = loja1
        ws.cell(row=linha, column=11).value = preco1
        ws.cell(row=linha, column=12).value = parcelamento1

        ws.cell(row=linha, column=15).value = loja2
        ws.cell(row=linha, column=16).value = preco2
        ws.cell(row=linha, column=17).value = parcelamento2

        ws.cell(row=linha, column=20).value = loja3
        ws.cell(row=linha, column=21).value = preco3
        ws.cell(row=linha, column=22).value = parcelamento3

        book.save(desktop)
        book.close()

        cont = cont + 1
        linha = linha + 1


    if btn_prod == 'PRODUTO NÃO LOCALIZADO':

        book = load_workbook(desktop)
        ws = book.worksheets[0]
        ws.cell(row=linha, column=7).value = hoje
        ws.cell(row=linha, column=8).value = 'Produto não encontrado'
        book.save(desktop)
        book.close()

        cont = cont + 1
        linha = linha + 1


browser.close()