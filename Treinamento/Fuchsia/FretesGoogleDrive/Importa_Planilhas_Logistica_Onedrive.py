from bs4 import BeautifulSoup
import os
import pandas as pd
import pyodbc
import sqlalchemy
from openpyxl import load_workbook

from datetime import timedelta, date
import datetime



hora_atual = str(datetime.datetime.today().time())
hora_atual = hora_atual[:5]

# Definir conexao 0 = PowerBiV2 / 1 = Datawarehouse
opcao = 0

configuracoes = pd.read_csv("C:/Python/database.cfg", header = 0, delimiter=";")

server = configuracoes['IP'].values[opcao]
database = configuracoes['Banco'].values[opcao]
username = configuracoes['Usuario'].values[opcao]
password = configuracoes['Senha'].values[opcao]


sql_string = 'mssql+pyodbc://' + username + ':' + password + '@' + server + '/' + database + '?driver=ODBC+Driver+13+for+SQL+Server'

cnxn = pyodbc.connect(
    'DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password, autocommit=True)
cursor = cnxn.cursor()

status_check_cursor = pyodbc.connect(
    'DRIVER={ODBC Driver 13 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password, autocommit=True)
cursor2 = status_check_cursor.cursor()


# Conexao com o servidor SQL
engine = sqlalchemy.create_engine(sql_string)
engine.connect()

# Data de hoje
# hoje = date.today().strftime('%d/%m/%Y')
hoje = date.today()
hojef = hoje.strftime('%d/%m/%Y')


# Busca os dados da planilha para pesquisar nos sites
desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'OneDrive - Mais Proxima\Importacao_Dados')
arquivo = '\LogES\coletas2020.xlsx'
desktop = desktop_path + arquivo


# IMPORTAÇAO PLANILHA DE COLETAS

# try:
#     dados = pd.read_excel(desktop, sheet_name='Coletas')
#     qtd_reg = str(len(dados))
#     del dados['OBS']
#     dados = dados.astype(str)
#
#     dados.to_sql(
#         name='Tabela_Onedrive_Coletas_Temp',  # database table name
#         con=engine,
#         if_exists='replace',
#         index=False,
#     )
#
#     # Grava na planilha que o registro foi importado para o SQL
#     book = load_workbook(desktop)
#     ws = book.worksheets[0]
#     lin = 2
#     col = 12
#     texto = 'Importado para SQL em ' + str(hojef) + ' - ' + hora_atual
#
#     for items in ws:
#         ws.cell(row=lin, column=col).value = texto
#
#         if lin <= len(dados):
#             lin = lin + 1
#     book.save(desktop)
#     book.close()
#     print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)
#
# except:
#     print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')


# # IMPORTAÇAO PLANILHA DE COTACAO
# arquivo = '\LogES\cotações2020.xlsx'
# desktop = desktop_path + arquivo
#
# try:
#     dados = pd.read_excel(desktop, sheet_name='Cotacao')
#     qtd_reg = str(len(dados))
#     del dados['OBS']
#     dados = dados.astype(str)
#
#     dados.to_sql(
#         name='Tabela_Onedrive_Cotacao_Temp',  # database table name
#         con=engine,
#         if_exists='replace',
#         index=False,
#     )
#
#     # Grava na planilha que o registro foi importado para o SQL
#     book = load_workbook(desktop)
#     ws = book.worksheets[0]
#     lin = 2
#     col = 9
#     texto = 'Importado para SQL em ' + str(hojef) + ' - ' + hora_atual
#
#     for items in ws:
#         ws.cell(row=lin, column=col).value = texto
#
#         if lin <= len(dados):
#             lin = lin + 1
#     book.save(desktop)
#     book.close()
#     print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)
#
#
# except:
#     print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')




# # IMPORTAÇAO PLANILHA NFS E ESTOQUE DIA
# arquivo = '\LogES\estoque_nfs_expedidas_dia.xlsx'
# desktop = desktop_path + arquivo
#
# try:
#     dados = pd.read_excel(desktop, sheet_name='Expedicao')
#     qtd_reg = str(len(dados))
#     del dados['OBS']
#     dados.fillna(0)
#     dados = dados.astype(str)
#
#     dados.to_sql(
#         name='Tabela_Onedrive_Expedicoes_Dia_Temp',  # database table name
#         con=engine,
#         if_exists='replace',
#         index=False,
#     )
#
#     # Grava na planilha que o registro foi importado para o SQL
#     book = load_workbook(desktop)
#     ws = book.worksheets[0]
#     lin = 2
#     col = 6
#     texto = 'Importado para SQL em ' + str(hojef) + ' - ' + hora_atual
#
#     for items in ws:
#         ws.cell(row=lin, column=col).value = texto
#
#         if lin <= len(dados):
#             lin = lin + 1
#     book.save(desktop)
#     book.close()
#     print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)
#
#
# except:
#     print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')




# # IMPORTAÇAO PLANILHA EXPEDICOES
# arquivo = '\LogES\expedições_2020.xlsx'
# desktop = desktop_path + arquivo
#
# try:
#     dados = pd.read_excel(desktop, sheet_name='Base_expedicoes')
#     qtd_reg = str(len(dados))
#     del dados['OBS']
#
#     dados = dados.astype(str)
#
#     dados.to_sql(
#         name='Tabela_Onedrive_Nfs_Expedidas_Temp',  # database table name
#         con=engine,
#         if_exists='replace',
#         index=False,
#     )
#
#     # Grava na planilha que o registro foi importado para o SQL
#     book = load_workbook(desktop)
#     ws = book.worksheets[0]
#     lin = 2
#     col = 8
#     texto = 'Importado para SQL em ' + str(hojef) + ' - ' + hora_atual
#
#     for items in ws:
#         ws.cell(row=lin, column=col).value = texto
#
#         if lin <= len(dados):
#             lin = lin + 1
#     book.save(desktop)
#     book.close()
#     print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)
#
#
# except:
#     print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')





# IMPORTAÇAO PLANILHA NFS PENDENTES DE EXPEDICAO
arquivo = '\LogES\pendentes_expedição.xlsx'
desktop = desktop_path + arquivo

try:
    dados = pd.read_excel(desktop, sheet_name='NF_Pendentes')
    qtd_reg = str(len(dados))
    del dados['OBS']
    dados = dados.astype(str)

    dados.to_sql(
        name='Tabela_Onedrive_NF_Pendentes_Temp',  # database table name
        con=engine,
        if_exists='replace',
        index=False,
    )

    # Grava na planilha que o registro foi importado para o SQL
    book = load_workbook(desktop)
    ws = book.worksheets[0]
    lin = 2
    col = 3
    texto = 'Importado para SQL em ' + str(hojef) + ' - ' + hora_atual

    for items in ws:
        ws.cell(row=lin, column=col).value = texto

        if lin <= len(dados):
            lin = lin + 1
    book.save(desktop)
    book.close()
    print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)


except:
    print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')



# # IMPORTAÇAO PLANILHA RECEBIMENTOS
# arquivo = '\LogES\Recebimentos.xlsx'
# desktop = desktop_path + arquivo
#
# try:
#     dados = pd.read_excel(desktop, sheet_name='Recebimentos')
#     dados = dados.fillna(None)
#     qtd_reg = str(len(dados))
#     del dados['OBS']
#     dados = dados.astype(str)
#
#
#     dados.to_sql(
#         name='Tabela_Onedrive_Recebimentos_Temp',  # database table name
#         con=engine,
#         if_exists='replace',
#         index=False,
#     )
#
#     # Grava na planilha que o registro foi importado para o SQL
#     book = load_workbook(desktop)
#     ws = book.worksheets[0]
#     lin = 2
#     col = 9
#     texto = 'Importado para SQL em ' + str(hojef) + ' - ' + hora_atual
#
#     for items in ws:
#         ws.cell(row=lin, column=col).value = texto
#
#         if lin <= len(dados):
#             lin = lin + 1
#     book.save(desktop)
#     book.close()
#     print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)
#
#
# except:
#     print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')


# # IMPORTAÇAO PLANILHA DE COTACAO
# arquivo = '\LogES\Faturas.xlsx'
# desktop = desktop_path + arquivo
#
# try:
#     dados = pd.read_excel(desktop, sheet_name='Faturas')
#     qtd_reg = str(len(dados))
#     del dados['OBS']
#     dados['Valor_Frete'] =  dados['Valor_Frete'].astype(int)
#     dados = dados.astype(str)
#
#     dados.to_sql(
#         name='Tabela_Onedrive_FaturasFrete_Temp',  # database table name
#         con=engine,
#         if_exists='replace',
#         index=False,
#     )
#
#     # Grava na planilha que o registro foi importado para o SQL
#     book = load_workbook(desktop)
#     ws = book.worksheets[0]
#     lin = 2
#     col = 5
#     texto = 'Importado para SQL em ' + str(hojef) + ' - ' + hora_atual
#
#     for items in ws:
#         ws.cell(row=lin, column=col).value = texto
#
#         if lin <= len(dados):
#             lin = lin + 1
#     book.save(desktop)
#     book.close()
#     print('A planilha ' + arquivo + ' foi importada com sucesso... Total de registros: ' + qtd_reg)
#
#
# except:
#     print('A planilha ' + arquivo + ' esta aberta, importacao nao realizada...')







# # Executando PROCEDURES no banco SQL
# print('')
# print('Executando Onedrive_Cotacao')
# cursor.execute("EXEC PowerBIv2..Onedrive_Cotacao")
# while 1:
#     q = status_check_cursor.execute("select RunningStatus from PowerBIv2..Acompanhamento_Rotinas_SQL_Python where Script_Procedure = 'Tabela_Onedrive_Cotacao'").fetchone()
#     if q[0] == 0:
#         break
#
# print('Executando Onedrive_Coletas')
# cursor.execute("EXEC PowerBIv2..Onedrive_Coletas")
# while 1:
#     q = status_check_cursor.execute("select RunningStatus from PowerBIv2..Acompanhamento_Rotinas_SQL_Python where Script_Procedure = 'Tabela_Onedrive_Coletas'").fetchone()
#     if q[0] == 0:
#         break
#
# print('Executando Onedrive_NFs_Expedidas')
# cursor.execute("EXEC PowerBIv2..Onedrive_NFs_Expedidas")
# while 1:
#     q = status_check_cursor.execute("select RunningStatus from PowerBIv2..Acompanhamento_Rotinas_SQL_Python where Script_Procedure = 'Tabela_Onedrive_Expedicoes_Dia'").fetchone()
#     if q[0] == 0:
#         break
#
#
print('Executando Onedrive_NF_Pendentes')
cursor.execute("EXEC PowerBIv2..Onedrive_NF_Pendentes")
while 1:
    q = status_check_cursor.execute("select RunningStatus from PowerBIv2..Acompanhamento_Rotinas_SQL_Python where Script_Procedure = 'Tabela_Onedrive_NF_Pendentes'").fetchone()
    if q[0] == 0:
        break
#
# print('Executando Onedrive_Expedicoes_Dia')
# cursor.execute("EXEC PowerBIv2..Onedrive_Expedicoes_Dia")
# while 1:
#     q = status_check_cursor.execute("select RunningStatus from PowerBIv2..Acompanhamento_Rotinas_SQL_Python where Script_Procedure = 'Tabela_Onedrive_Nfs_Expedidas'").fetchone()
#     if q[0] == 0:
#         break
#
# print('Executando Onedrive_Recebimentos')
# cursor.execute("EXEC PowerBIv2..Onedrive_Recebimentos")
# while 1:
#     q = status_check_cursor.execute("select RunningStatus from PowerBIv2..Acompanhamento_Rotinas_SQL_Python where Script_Procedure = 'Tabela_Onedrive_Recebimentos'").fetchone()
#     if q[0] == 0:
#         break
#
#
# print('Executando Onedrive_Faturas_Frete')
# cursor.execute("EXEC PowerBIv2..Onedrive_Faturas_Frete")
# while 1:
#     q = status_check_cursor.execute("select RunningStatus from PowerBIv2..Acompanhamento_Rotinas_SQL_Python where Script_Procedure = 'Tabela_Onedrive_Faturas_Frete'").fetchone()
#     if q[0] == 0:
#         break


print('Importacoes Finalizadas')
cursor.close()
cursor2.close()
